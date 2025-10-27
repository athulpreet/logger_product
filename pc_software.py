import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import csv
import serial
import serial.tools.list_ports
import threading
import time
import json
from datetime import datetime

# For Excel export
from openpyxl import Workbook

# For PDF export
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch

# For Plotting
import matplotlib
matplotlib.use("TkAgg")  # use TkAgg backend for tkinter
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import NavigationToolbar2Tk

# ---------------------------------------------------
# Global constants for commands (strings/bytes):
# ---------------------------------------------------
BAUD_RATE = 115200
LOGGER_CMD = b'B=GET_72HRS_DATA\n'

CMD_HEAD = b'C=HEAD\n'
CMD_TRP = b'D=TRP\n'
CMD_1TP = b'E=1TP\n'
CMD_6TP = b'F=6TP\n'
CMD_VIOL = b'G=VIOL\n'
CMD_RST = b'H=RST\n'
CMD_TP = b'I=TP\n'  

SET_SPEED_PREFIX = "SET_SPEEDI="
SET_LIMP_PREFIX  = "SET_LIMPJ="
SET_TIME_PREFIX  = "SET_TIMEK="

# Flag constants for print/download
FLAG_PRINT = "P"    # Print to printer
FLAG_DOWNLOAD = "D" # Download to PC

SERIAL_READ_TIMEOUT = 2.0

class HeaderDetailsWindow:
    def __init__(self, parent):
        self.parent = parent
        self.window = tk.Toplevel(parent.root)
        self.window.title("Vehicle Header Details")
        self.window.geometry("500x600")
        self.window.grab_set()  # Make window modal
        
        # Center the window
        self.window.transient(parent.root)
        self.window.geometry("+%d+%d" % (parent.root.winfo_rootx() + 50, parent.root.winfo_rooty() + 50))
        
        # Initialize header data
        self.header_data = {
            'vehicle_owner_name': '',
            'vehicle_owner_id': '',
            'owner_phone_no': '',
            'vehicle_reg_number': '',
            'chassis_no': '',
            'vehicle_make_type': '',
            'certificate_number': '',
            'limiter_type': '',
            'limiter_serial': '',
            'date_of_fitting': '',
            'fitting_agent_name': '',
            'fitting_agent_id': '',
            'name_location_station': '',
            'email_address': '',
            'agent_phone_number': '',
            'business_reg_no': ''
        }
        
        # Set up a protocol to handle window closing
        self.window.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        self.create_widgets()
        
    def on_closing(self):
        """Clean up when window is closed"""
        # Unbind the mousewheel event
        self.window.unbind_all("<MouseWheel>")
        # Destroy the window
        self.window.destroy()
        
    def create_widgets(self):
        # Main frame with scrollbar
        main_frame = ttk.Frame(self.window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Title
        title_label = ttk.Label(main_frame, text="Vehicle Header Details", 
                                font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 20))
        
        # Create scrollable frame
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Entry fields
        self.entries = {}
        
        # Vehicle Owner Information
        owner_frame = ttk.LabelFrame(scrollable_frame, text="Vehicle Owner Information", padding="10")
        owner_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(owner_frame, text="Vehicle Owner Name:").grid(row=0, column=0, sticky="w", pady=2)
        self.entries['vehicle_owner_name'] = ttk.Entry(owner_frame, width=40)
        self.entries['vehicle_owner_name'].grid(row=0, column=1, pady=2, padx=(10, 0))
        
        ttk.Label(owner_frame, text="Vehicle Owner ID:").grid(row=1, column=0, sticky="w", pady=2)
        self.entries['vehicle_owner_id'] = ttk.Entry(owner_frame, width=40)
        self.entries['vehicle_owner_id'].grid(row=1, column=1, pady=2, padx=(10, 0))
        
        ttk.Label(owner_frame, text="Owner Phone Number:").grid(row=2, column=0, sticky="w", pady=2)
        self.entries['owner_phone_no'] = ttk.Entry(owner_frame, width=40)
        self.entries['owner_phone_no'].grid(row=2, column=1, pady=2, padx=(10, 0))
        
        ttk.Label(owner_frame, text="Email Address:").grid(row=3, column=0, sticky="w", pady=2)
        self.entries['email_address'] = ttk.Entry(owner_frame, width=40)
        self.entries['email_address'].grid(row=3, column=1, pady=2, padx=(10, 0))
        
        # Vehicle Information
        vehicle_frame = ttk.LabelFrame(scrollable_frame, text="Vehicle Information", padding="10")
        vehicle_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(vehicle_frame, text="Vehicle Registration Number:").grid(row=0, column=0, sticky="w", pady=2)
        self.entries['vehicle_reg_number'] = ttk.Entry(vehicle_frame, width=40)
        self.entries['vehicle_reg_number'].grid(row=0, column=1, pady=2, padx=(10, 0))
        
        ttk.Label(vehicle_frame, text="Chassis Number:").grid(row=1, column=0, sticky="w", pady=2)
        self.entries['chassis_no'] = ttk.Entry(vehicle_frame, width=40)
        self.entries['chassis_no'].grid(row=1, column=1, pady=2, padx=(10, 0))
        
        ttk.Label(vehicle_frame, text="Vehicle Make & Type:").grid(row=2, column=0, sticky="w", pady=2)
        self.entries['vehicle_make_type'] = ttk.Entry(vehicle_frame, width=40)
        self.entries['vehicle_make_type'].grid(row=2, column=1, pady=2, padx=(10, 0))
        
        # Speed Limiter Information
        limiter_frame = ttk.LabelFrame(scrollable_frame, text="Speed Limiter Information", padding="10")
        limiter_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(limiter_frame, text="Certificate Number:").grid(row=0, column=0, sticky="w", pady=2)
        self.entries['certificate_number'] = ttk.Entry(limiter_frame, width=40)
        self.entries['certificate_number'].grid(row=0, column=1, pady=2, padx=(10, 0))
        
        ttk.Label(limiter_frame, text="Limiter Type:").grid(row=1, column=0, sticky="w", pady=2)
        self.entries['limiter_type'] = ttk.Entry(limiter_frame, width=40)
        self.entries['limiter_type'].grid(row=1, column=1, pady=2, padx=(10, 0))
        
        ttk.Label(limiter_frame, text="Limiter Serial:").grid(row=2, column=0, sticky="w", pady=2)
        self.entries['limiter_serial'] = ttk.Entry(limiter_frame, width=40)
        self.entries['limiter_serial'].grid(row=2, column=1, pady=2, padx=(10, 0))
        
        ttk.Label(limiter_frame, text="Date of Fitting:").grid(row=3, column=0, sticky="w", pady=2)
        self.entries['date_of_fitting'] = ttk.Entry(limiter_frame, width=40)
        self.entries['date_of_fitting'].grid(row=3, column=1, pady=2, padx=(10, 0))
        self.entries['date_of_fitting'].insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        # Fitting Agent Information
        agent_frame = ttk.LabelFrame(scrollable_frame, text="Fitting Agent Information", padding="10")
        agent_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(agent_frame, text="Fitting Agent Name:").grid(row=0, column=0, sticky="w", pady=2)
        self.entries['fitting_agent_name'] = ttk.Entry(agent_frame, width=40)
        self.entries['fitting_agent_name'].grid(row=0, column=1, pady=2, padx=(10, 0))
        
        ttk.Label(agent_frame, text="Fitting Agent ID:").grid(row=1, column=0, sticky="w", pady=2)
        self.entries['fitting_agent_id'] = ttk.Entry(agent_frame, width=40)
        self.entries['fitting_agent_id'].grid(row=1, column=1, pady=2, padx=(10, 0))
        
        ttk.Label(agent_frame, text="Name & Location of Station:").grid(row=2, column=0, sticky="w", pady=2)
        self.entries['name_location_station'] = ttk.Entry(agent_frame, width=40)
        self.entries['name_location_station'].grid(row=2, column=1, pady=2, padx=(10, 0))
        
        ttk.Label(agent_frame, text="Agent Phone Number:").grid(row=3, column=0, sticky="w", pady=2)
        self.entries['agent_phone_number'] = ttk.Entry(agent_frame, width=40)
        self.entries['agent_phone_number'].grid(row=3, column=1, pady=2, padx=(10, 0))
        
        ttk.Label(agent_frame, text="Business Registration No:").grid(row=4, column=0, sticky="w", pady=2)
        self.entries['business_reg_no'] = ttk.Entry(agent_frame, width=40)
        self.entries['business_reg_no'].grid(row=4, column=1, pady=2, padx=(10, 0))
        
        # Buttons
        button_frame = ttk.Frame(scrollable_frame)
        button_frame.pack(fill=tk.X, pady=20)
        
        ttk.Button(button_frame, text="Send Header Data", 
                   command=self.send_header_data).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Save Header Data", 
                   command=self.save_header_data).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Load Header Data", 
                   command=self.load_header_data).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Clear All", 
                   command=self.clear_all_fields).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Close", 
                   command=self.on_closing).pack(side=tk.RIGHT)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Bind mousewheel to canvas
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        self.window.bind_all("<MouseWheel>", _on_mousewheel)  # Store binding reference
        
    def send_header_data(self):
        """
        Send header data to the device.
        MODIFIED: Sends data starting with '*', ending with '#', in 64-byte packets.
        MODIFIED: Final packet is NOT padded with null bytes.
        """
        if not self.parent.serial_port:
            messagebox.showerror("No Port", "Please connect to a port first.")
            return
            
        # Collect all header data
        header_data = {}
        missing_fields = []
        
        for key, entry in self.entries.items():
            value = entry.get().strip()
            header_data[key] = value
            if not value:
                missing_fields.append(key.replace('_', ' ').title())
        
        if missing_fields:
            result = messagebox.askyesno("Missing Fields", 
                                       f"The following fields are empty:\n{', '.join(missing_fields)}\n\n"
                                       "Do you want to continue sending the header data?")
            if not result:
                return
        
        # --- START OF MODIFIED SEND LOGIC ---
        
        # 1. Format header data according to KS 2295-1:2023 standard
        header_string = self.format_header_data(header_data)
        
        # 2. Add start and end markers
        full_data_string = f"*{header_string}#"
        
        # 3. Convert to bytes
        full_data_bytes = full_data_string.encode('utf-8')
        
        # 4. Define packet size
        PACKET_SIZE = 64
        
        try:
            with serial.Serial(self.parent.serial_port, BAUD_RATE, timeout=SERIAL_READ_TIMEOUT) as ser:
                
                self.parent.log_text.insert(tk.END, f"\n--- Sending Header Data in {PACKET_SIZE}-byte (max) packets ---\n")
                
                # 5. Loop and send in packets
                for i in range(0, len(full_data_bytes), PACKET_SIZE):
                    # Get the chunk (will be 64 bytes or less for the final packet)
                    packet = full_data_bytes[i : i + PACKET_SIZE]
                    
                    # --- PADDING LOGIC REMOVED ---
                    # The last packet will be sent as-is, without null byte padding.
                        
                    # Send the packet (64 bytes or less)
                    ser.write(packet)
                    
                    # Log what was sent (using repr() to show non-printable chars)
                    self.parent.log_text.insert(tk.END, f"Sent packet {i//PACKET_SIZE + 1} ({len(packet)} bytes): {repr(packet)}\n")
                    self.parent.log_text.see(tk.END)
                    
                    # Small delay between packets for device processing
                    time.sleep(0.05) # 50ms delay
                
                # --- End of packet sending loop ---
                
                # Wait for device to process all packets
                time.sleep(1)  
                
                # Read response
                responses = []
                while ser.in_waiting:
                    line = ser.readline().decode('utf-8', errors='ignore').strip()
                    if line:
                        responses.append(line)
                
                # Display results
                self.parent.log_text.insert(tk.END, f"\n--- Header Data Sent ---\n")
                self.parent.log_text.insert(tk.END, f"Total bytes sent: {len(full_data_bytes)}\n")
                
                if responses:
                    self.parent.log_text.insert(tk.END, "Device Response:\n")
                    for resp in responses:
                        self.parent.log_text.insert(tk.END, resp + "\n")
                    
                    # Save the responses to log_data for export
                    raw_data = responses.copy()
                    self.parent.process_log_data(raw_data)
                else:
                    self.parent.log_text.insert(tk.END, "(No response from device)\n")
                
                self.parent.log_text.see(tk.END)
                messagebox.showinfo("Success", "Header data sent successfully!")
                
        except serial.SerialException as e:
            messagebox.showerror("Serial Error", f"Could not send header data: {e}")
        except Exception as ex:
            messagebox.showerror("Error", f"An error occurred: {ex}")
        
        # --- END OF MODIFIED SEND LOGIC ---

    def format_header_data(self, data):
        """Format header data according to KS 2295-1:2023 Annex B.1.2"""
        # This function is UNCHANGED. It still creates a newline-separated string.
        # The start/end markers (*) (#) are added in send_header_data.
        header_lines = []
        
        header_lines.append(f"A={data.get('vehicle_owner_name', '')}")
        header_lines.append(f"{data.get('vehicle_owner_id', '')}")
        header_lines.append(f"{data.get('owner_phone_no', '')}")
        header_lines.append(f"{data.get('vehicle_reg_number', '')}")
        header_lines.append(f"{data.get('chassis_no', '')}")
        header_lines.append(f"{data.get('vehicle_make_type', '')}")
        header_lines.append(f"{data.get('certificate_number', '')}")
        header_lines.append(f"{data.get('limiter_type', '')}")
        header_lines.append(f"{data.get('limiter_serial', '')}")
        header_lines.append(f"{data.get('date_of_fitting', '')}")
        header_lines.append(f"{data.get('fitting_agent_name', '')}")
        header_lines.append(f"{data.get('fitting_agent_id', '')}")
        header_lines.append(f"{data.get('name_location_station', '')}")
        header_lines.append(f"{data.get('email_address', '')}")
        header_lines.append(f"{data.get('agent_phone_number', '')}")
        header_lines.append(f"{data.get('business_reg_no', '')}")
        
        # The final '\n' here is correct, it will be part of the data sent.
        return '\n'.join(header_lines) + '\n'
    
    def save_header_data(self):
        """Save header data to a JSON file"""
        header_data = {}
        for key, entry in self.entries.items():
            header_data[key] = entry.get().strip()
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON Files", "*.json"), ("All Files", "*.*")],
            title="Save Header Data"
        )
        
        if file_path:
            try:
                with open(file_path, 'w') as f:
                    json.dump(header_data, f, indent=4)
                messagebox.showinfo("Success", f"Header data saved to:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Could not save header data: {e}")
    
    def load_header_data(self):
        """Load header data from a JSON file"""
        file_path = filedialog.askopenfilename(
            filetypes=[("JSON Files", "*.json"), ("All Files", "*.*")],
            title="Load Header Data"
        )
        
        if file_path:
            try:
                with open(file_path, 'r') as f:
                    header_data = json.load(f)
                
                for key, value in header_data.items():
                    if key in self.entries:
                        self.entries[key].delete(0, tk.END)
                        self.entries[key].insert(0, value)
                
                messagebox.showinfo("Success", f"Header data loaded from:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Could not load header data: {e}")
    
    def clear_all_fields(self):
        """Clear all entry fields"""
        result = messagebox.askyesno("Clear All", "Are you sure you want to clear all fields?")
        if result:
            for entry in self.entries.values():
                entry.delete(0, tk.END)

class LogDownloaderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Vehicle Limiter Control & Log Downloader")

        # Data containers
        self.log_data = []
        self.is_downloading = False
        self.serial_port = None  # Will store user-selected COM port or None if not connected
        self.time_update_running = False  # Flag to control time update thread
        self.output_mode = tk.StringVar(value=FLAG_DOWNLOAD)  # Default to download mode

        # ----- Main Frame -----
        main_frame = ttk.Frame(self.root, padding="10 10 10 10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        style = ttk.Style()
        style.theme_use("clam")

        # ------------------------------------------------
        # 1. Connection Frame: Select/Refresh/Connect
        # ------------------------------------------------
        conn_frame = ttk.Labelframe(main_frame, text="Connection", padding="10 10 10 10")
        conn_frame.pack(fill=tk.X, expand=False, pady=5)

        # A dropdown (combobox) to list available ports
        ttk.Label(conn_frame, text="Select Port:").grid(row=0, column=0, padx=5, pady=5, sticky="e")

        self.port_var = tk.StringVar()  # holds the currently selected port from combobox
        self.port_combo = ttk.Combobox(conn_frame, textvariable=self.port_var, state="readonly", width=15)
        self.port_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # Button to refresh port list
        self.refresh_button = ttk.Button(conn_frame, text="Refresh Ports", command=self.refresh_ports)
        self.refresh_button.grid(row=0, column=2, padx=5, pady=5)

        # Connect button
        self.connect_button = ttk.Button(conn_frame, text="Connect", command=self.on_connect)
        self.connect_button.grid(row=0, column=3, padx=5, pady=5)
        
        # Output Mode (Print/Download) toggle
        ttk.Label(conn_frame, text="Output Mode:").grid(row=0, column=4, padx=(15, 5), pady=5, sticky="e")
        ttk.Radiobutton(conn_frame, text="Print", 
                        variable=self.output_mode, value=FLAG_PRINT).grid(row=0, column=5, padx=5, pady=5)
        ttk.Radiobutton(conn_frame, text="Download", 
                        variable=self.output_mode, value=FLAG_DOWNLOAD).grid(row=0, column=6, padx=5, pady=5)

        # ------------------------------------------------
        # 2. Header Management Frame
        # ------------------------------------------------
        header_frame = ttk.Labelframe(main_frame, text="Header Management", padding="10 10 10 10")
        header_frame.pack(fill=tk.X, expand=False, pady=5)

        self.header_button = ttk.Button(header_frame, text="Configure Header Details", command=self.open_header_window)
        self.header_button.grid(row=0, column=0, padx=5, pady=5)

        # ------------------------------------------------
        # 3. Download & Export Frame
        # ------------------------------------------------
        top_frame = ttk.Labelframe(main_frame, text="Download & Export", padding="10 10 10 10")
        top_frame.pack(fill=tk.X, expand=False, pady=5)

        self.download_button = ttk.Button(top_frame, text="Download (GET_72HRS_DATA)", command=self.initiate_download)
        self.download_button.grid(row=0, column=0, padx=5, pady=5)

        self.csv_button = ttk.Button(top_frame, text="Export CSV", command=self.export_csv)
        self.csv_button.grid(row=0, column=1, padx=5, pady=5)

        self.excel_button = ttk.Button(top_frame, text="Export Excel", command=self.export_excel)
        self.excel_button.grid(row=0, column=2, padx=5, pady=5)

        self.pdf_button = ttk.Button(top_frame, text="Export PDF", command=self.export_pdf)
        self.pdf_button.grid(row=0, column=3, padx=5, pady=5)

        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(top_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=1, column=0, columnspan=4, sticky="ew", pady=5)

        # ------------------------------------------------
        # 4. Commands & Settings Frame
        # ------------------------------------------------
        cmd_frame = ttk.Labelframe(main_frame, text="Commands & Settings", padding="10 10 10 10")
        cmd_frame.pack(fill=tk.X, expand=False, pady=5)

        self.head_button = ttk.Button(cmd_frame, text="HEAD", command=lambda: self.send_simple_command(CMD_HEAD))
        self.head_button.grid(row=0, column=0, padx=5, pady=5)

        self.trp_button = ttk.Button(cmd_frame, text="TRP", command=lambda: self.send_simple_command(CMD_TRP))
        self.trp_button.grid(row=0, column=1, padx=5, pady=5)

        self.tp1_button = ttk.Button(cmd_frame, text="1TP", command=lambda: self.send_simple_command(CMD_1TP))
        self.tp1_button.grid(row=0, column=2, padx=5, pady=5)

        self.tp6_button = ttk.Button(cmd_frame, text="6TP", command=lambda: self.send_simple_command(CMD_6TP))
        self.tp6_button.grid(row=0, column=3, padx=5, pady=5)

        self.viol_button = ttk.Button(cmd_frame, text="VIOL", command=lambda: self.send_simple_command(CMD_VIOL))
        self.viol_button.grid(row=0, column=4, padx=5, pady=5)

        self.rst_button = ttk.Button(cmd_frame, text="RST", command=lambda: self.send_simple_command(CMD_RST))
        self.rst_button.grid(row=0, column=5, padx=5, pady=5)
        
        # Add the new TP button
        self.tp_button = ttk.Button(cmd_frame, text="TP", command=lambda: self.send_simple_command(CMD_TP))
        self.tp_button.grid(row=0, column=6, padx=5, pady=5)

        ttk.Label(cmd_frame, text="Set Speed (km/h):").grid(row=1, column=0, sticky="e", padx=5)
        self.set_speed_entry = ttk.Entry(cmd_frame, width=8)
        self.set_speed_entry.grid(row=1, column=1, sticky="w", padx=5)
        self.set_speed_button = ttk.Button(cmd_frame, text="Send", command=self.send_set_speed)
        self.set_speed_button.grid(row=1, column=2, padx=5, pady=5)

        ttk.Label(cmd_frame, text="Limp Speed (km/h):").grid(row=2, column=0, sticky="e", padx=5)
        self.set_limp_entry = ttk.Entry(cmd_frame, width=8)
        self.set_limp_entry.grid(row=2, column=1, sticky="w", padx=5)
        self.set_limp_button = ttk.Button(cmd_frame, text="Send", command=self.send_limp_speed)
        self.set_limp_button.grid(row=2, column=2, padx=5, pady=5)

        # Time setting - auto-updating entry field
        ttk.Label(cmd_frame, text="Set Device Time:").grid(row=3, column=0, sticky="e", padx=5)
        self.set_time_var = tk.StringVar()  # Variable to hold the current time
        self.set_time_entry = ttk.Entry(cmd_frame, width=18, textvariable=self.set_time_var)
        self.set_time_entry.grid(row=3, column=1, sticky="w", padx=5)
        self.set_time_button = ttk.Button(cmd_frame, text="Send", command=self.send_time_setting)
        self.set_time_button.grid(row=3, column=2, padx=5, pady=5)

        # Update the time initially
        self.update_time_display()
        
        # Start the time update thread
        self.time_update_running = True
        self.time_thread = threading.Thread(target=self.time_update_loop, daemon=True)
        self.time_thread.start()

        # ------------------------------------------------
        # 5. Filters & Visualization
        # ------------------------------------------------
        middle_frame = ttk.Labelframe(main_frame, text="Data Filters & Visualization", padding="10 10 10 10")
        middle_frame.pack(fill=tk.X, expand=False, pady=5)

        ttk.Label(middle_frame, text="Filter by Date (YYMMDD):").grid(row=0, column=0, padx=5, pady=2, sticky="e")
        self.filter_date_entry = ttk.Entry(middle_frame, width=10)
        self.filter_date_entry.grid(row=0, column=1, padx=5, pady=2, sticky="w")

        ttk.Label(middle_frame, text="Filter by Time (HHMMSS):").grid(row=1, column=0, padx=5, pady=2, sticky="e")
        self.filter_time_entry = ttk.Entry(middle_frame, width=10)
        self.filter_time_entry.grid(row=1, column=1, padx=5, pady=2, sticky="w")

        self.filter_button = ttk.Button(middle_frame, text="Apply Filter", command=self.apply_filter)
        self.filter_button.grid(row=0, column=2, rowspan=2, padx=10, pady=2, sticky="ns")

        self.plot_button = ttk.Button(middle_frame, text="Plot Speed vs Time", command=self.plot_speed_vs_time)
        self.plot_button.grid(row=0, column=3, rowspan=2, padx=10, pady=2)

        # ------------------------------------------------
        # 6. Bottom Frame: Log Text (Made Much Larger)
        # ------------------------------------------------
        bottom_frame = ttk.Labelframe(main_frame, text="Log Data & Terminal Output", padding="10 10 10 10")
        bottom_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        # Create a frame for the text widget and scrollbars
        text_frame = ttk.Frame(bottom_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)

        # Create the text widget with larger dimensions
        self.log_text = tk.Text(text_frame, 
                                width=100, 
                                height=25,  # Increased from 15 to 25
                                wrap="none",
                                font=("Consolas", 10),  # Monospace font for better readability
                                bg="black",
                                fg="light green",
                                insertbackground="white")
        
        # Create both vertical and horizontal scrollbars
        v_scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        h_scrollbar = ttk.Scrollbar(text_frame, orient=tk.HORIZONTAL, command=self.log_text.xview)
        
        # Configure text widget scrollbars
        self.log_text.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack scrollbars and text widget
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Add a clear log button
        log_controls_frame = ttk.Frame(bottom_frame)
        log_controls_frame.pack(fill=tk.X, pady=(5, 0))
        
        ttk.Button(log_controls_frame, text="Clear Log", 
                   command=self.clear_log).pack(side=tk.LEFT, padx=5)
        ttk.Button(log_controls_frame, text="Save Log to File", 
                   command=self.save_log).pack(side=tk.LEFT, padx=5)
                       
        # Set protocol for window closing
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Now that log_text is created, populate the port dropdown
        self.refresh_ports()
        
    def update_time_display(self):
        """Update the time display with the current time"""
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.set_time_var.set(current_time)
        
    def time_update_loop(self):
        """Loop that updates the time display every second"""
        while self.time_update_running:
            try:
                # Use after() to update the time display from the main thread
                self.root.after(0, self.update_time_display)
                time.sleep(1)  # Update once per second
            except Exception as e:
                print(f"Error in time update loop: {e}")
                # If there's an error, wait a bit longer before retrying
                time.sleep(5)
                
    def on_closing(self):
        """Handle application closing"""
        self.time_update_running = False
        # Wait a moment for the thread to terminate
        if hasattr(self, 'time_thread') and self.time_thread.is_alive():
            self.time_thread.join(timeout=0.1)
        self.root.destroy()
        
    def refresh_ports(self):
        """Refresh the list of available serial ports"""
        ports = [port.device for port in serial.tools.list_ports.comports()]
        self.port_combo['values'] = ports
        
        if ports:
            self.port_combo.current(0)  # Select first port
            self.log_text.insert(tk.END, f"Detected serial ports: {', '.join(ports)}\n")
        else:
            self.log_text.insert(tk.END, "No serial ports detected.\n")
        self.log_text.see(tk.END)
        
    def on_connect(self):
        """Connect to the selected serial port"""
        selected_port = self.port_var.get()
        
        if not selected_port:
            messagebox.showerror("No Port Selected", "Please select a port to connect to.")
            return
            
        try:
            # Test connection
            with serial.Serial(selected_port, BAUD_RATE, timeout=SERIAL_READ_TIMEOUT) as ser:
                self.serial_port = selected_port
                self.log_text.insert(tk.END, f"\nConnected to {selected_port} at {BAUD_RATE} baud.\n")
                self.log_text.see(tk.END)
                
                # Update button text
                self.connect_button.config(text="Connected")
                
                # Get device identification info
                self.log_text.insert(tk.END, "Requesting device information...\n")
                self.log_text.see(tk.END)
                
                # Wait for any initial data
                time.sleep(0.5)
                
                # Read any initial response
                responses = []
                while ser.in_waiting:
                    line = ser.readline().decode('utf-8', errors='ignore').strip()
                    if line:
                        responses.append(line)
                        self.log_text.insert(tk.END, f"Device: {line}\n")
                        self.log_text.see(tk.END)
                
                # Save any initial responses to log_data
                if responses:
                    self.process_log_data(responses)
                
        except serial.SerialException as e:
            messagebox.showerror("Connection Error", f"Could not connect to {selected_port}: {e}")
            self.log_text.insert(tk.END, f"Connection error: {e}\n")
            self.log_text.see(tk.END)
            
    def open_header_window(self):
        """Open header details configuration window"""
        # Clear log data when opening header window
        self.log_data = []
        HeaderDetailsWindow(self)
        
    def send_simple_command(self, command):
        """Send a simple command to the device"""
        if not self.serial_port:
            messagebox.showerror("No Port", "Please connect to a port first.")
            return
            
        try:
            # Clear log data before processing new command
            self.log_data = []
            
            with serial.Serial(self.serial_port, BAUD_RATE, timeout=SERIAL_READ_TIMEOUT) as ser:
                # Get the output flag
                output_flag = self.output_mode.get()
                
                # Decode the command for manipulation
                cmd_decoded = command.decode('utf-8', errors='ignore').strip()
                
                # Add the flag to the command
                flagged_command = f"{output_flag}:{cmd_decoded}\n".encode('utf-8')
                
                # Send command with flag
                ser.write(flagged_command)
                
                # Command name for logging
                self.log_text.insert(tk.END, f"\n--- Sent command: {cmd_decoded} ---\n")
                self.log_text.insert(tk.END, f"Output Mode: {'Print to Printer' if output_flag == FLAG_PRINT else 'Download to PC'}\n")
                
                # Wait for processing
                time.sleep(0.5)
                
                # Read response
                responses = []
                while ser.in_waiting:
                    line = ser.readline().decode('utf-8', errors='ignore').strip()
                    if line:
                        responses.append(line)
                
                if responses:
                    self.log_text.insert(tk.END, "Device Response:\n")
                    for resp in responses:
                        self.log_text.insert(tk.END, resp + "\n")
                    
                    # Save the responses to log_data for export
                    raw_data = responses.copy()
                    
                    # Flag if this is a HEAD command
                    is_head_command = (command == CMD_HEAD)
                    
                    # Process the data with command type info
                    self.process_log_data(raw_data, is_head_command=is_head_command)
                else:
                    self.log_text.insert(tk.END, "(No response from device)\n")
                
                self.log_text.see(tk.END)
                
        except serial.SerialException as e:
            messagebox.showerror("Serial Error", f"Could not send command: {e}")
            
    def send_set_speed(self):
        """Send speed setting to the device"""
        if not self.serial_port:
            messagebox.showerror("No Port", "Please connect to a port first.")
            return
            
        try:
            # Clear log data before processing new command
            self.log_data = []
            
            speed = self.set_speed_entry.get().strip()
            
            # Validate speed value
            try:
                speed_val = int(speed)
                if speed_val < 0 or speed_val > 250:  # Reasonable range
                    messagebox.showerror("Invalid Value", "Speed must be between 0 and 250 km/h.")
                    return
            except ValueError:
                messagebox.showerror("Invalid Value", "Please enter a valid speed value.")
                return
            
            # Format command (no flag for speed settings)
            command = f"{SET_SPEED_PREFIX}{speed}\n"
            
            with serial.Serial(self.serial_port, BAUD_RATE, timeout=SERIAL_READ_TIMEOUT) as ser:
                # Send command
                ser.write(command.encode('utf-8'))
                
                self.log_text.insert(tk.END, f"\n--- Set Speed Command ---\n")
                self.log_text.insert(tk.END, f"Command: {command}\n")
                
                # Wait for processing
                time.sleep(0.5)
                
                # Read response
                responses = []
                while ser.in_waiting:
                    line = ser.readline().decode('utf-8', errors='ignore').strip()
                    if line:
                        responses.append(line)
                
                if responses:
                    self.log_text.insert(tk.END, "Device Response:\n")
                    for resp in responses:
                        self.log_text.insert(tk.END, resp + "\n")
                    
                    # Save the responses to log_data for export
                    raw_data = responses.copy()
                    self.process_log_data(raw_data)
                else:
                    self.log_text.insert(tk.END, "(No response from device)\n")
                
                self.log_text.see(tk.END)
                
        except serial.SerialException as e:
            messagebox.showerror("Serial Error", f"Could not send speed setting: {e}")
            
    def send_limp_speed(self):
        """Send limp mode speed setting to the device"""
        if not self.serial_port:
            messagebox.showerror("No Port", "Please connect to a port first.")
            return
            
        try:
            # Clear log data before processing new command
            self.log_data = []
            
            speed = self.set_limp_entry.get().strip()
            
            # Validate speed value
            try:
                speed_val = int(speed)
                if speed_val < 0 or speed_val > 250:  # Reasonable range
                    messagebox.showerror("Invalid Value", "Limp speed must be between 0 and 250 km/h.")
                    return
            except ValueError:
                messagebox.showerror("Invalid Value", "Please enter a valid speed value.")
                return
            
            # Format command (no flag for limp speed settings)
            command = f"{SET_LIMP_PREFIX}{speed}\n"
            
            with serial.Serial(self.serial_port, BAUD_RATE, timeout=SERIAL_READ_TIMEOUT) as ser:
                # Send command
                ser.write(command.encode('utf-8'))
                
                self.log_text.insert(tk.END, f"\n--- Set Limp Speed Command ---\n")
                self.log_text.insert(tk.END, f"Command: {command}\n")
                
                # Wait for processing
                time.sleep(0.5)
                
                # Read response
                responses = []
                while ser.in_waiting:
                    line = ser.readline().decode('utf-8', errors='ignore').strip()
                    if line:
                        responses.append(line)
                
                if responses:
                    self.log_text.insert(tk.END, "Device Response:\n")
                    for resp in responses:
                        self.log_text.insert(tk.END, resp + "\n")
                    
                    # Save the responses to log_data for export
                    raw_data = responses.copy()
                    self.process_log_data(raw_data)
                else:
                    self.log_text.insert(tk.END, "(No response from device)\n")
                
                self.log_text.see(tk.END)
                
        except serial.SerialException as e:
            messagebox.showerror("Serial Error", f"Could not send limp speed setting: {e}")
            
    def send_time_setting(self):
        """Send the current time to the device"""
        if not self.serial_port:
            messagebox.showerror("No Port", "Please connect to a port first.")
            return
        
        try:
            # Clear log data before processing new command
            self.log_data = []
            
            # Get the current time in the format expected by the device
            current_time = datetime.now().strftime("%y%m%d%H%M%S")  # Format as YYMMDDHHMMSS
            command = f"{SET_TIME_PREFIX}{current_time}\n"
            
            with serial.Serial(self.serial_port, BAUD_RATE, timeout=SERIAL_READ_TIMEOUT) as ser:
                ser.write(command.encode('utf-8'))
                time.sleep(0.5)  # Wait for device to process
                
                # Read response
                responses = []
                while ser.in_waiting:
                    line = ser.readline().decode('utf-8', errors='ignore').strip()
                    if line:
                        responses.append(line)
                
                # Display results
                self.log_text.insert(tk.END, f"\n--- Time Setting Sent ---\n")
                self.log_text.insert(tk.END, f"Command: {command}\n")
                
                if responses:
                    self.log_text.insert(tk.END, "Device Response:\n")
                    for resp in responses:
                        self.log_text.insert(tk.END, resp + "\n")
                    
                    # Save the responses to log_data for export
                    raw_data = responses.copy()
                    self.process_log_data(raw_data)
                else:
                    self.log_text.insert(tk.END, "(No response from device)\n")
                
                self.log_text.see(tk.END)
                
        except serial.SerialException as e:
            messagebox.showerror("Serial Error", f"Could not send time setting: {e}")
        except Exception as ex:
            messagebox.showerror("Error", f"An error occurred: {ex}")
            
    def initiate_download(self):
        """Start downloading log data from the device"""
        if not self.serial_port:
            messagebox.showerror("No Port", "Please connect to a port first.")
            return
            
        if self.is_downloading:
            messagebox.showinfo("Download in Progress", "Download already in progress.")
            return
            
        # Reset log data and progress
        self.log_data = []
        self.progress_var.set(0)
        self.is_downloading = True
        
        # Start the download thread
        download_thread = threading.Thread(target=self.download_thread, daemon=True)
        download_thread.start()
        
    def download_thread(self):
        """Thread function for downloading log data"""
        try:
            with serial.Serial(self.serial_port, BAUD_RATE, timeout=SERIAL_READ_TIMEOUT) as ser:
                # Get the output flag
                output_flag = self.output_mode.get()
                
                # Send download command with flag
                self.log_text.insert(tk.END, "\n--- Starting Log Download ---\n")
                self.log_text.insert(tk.END, f"Output Mode: {'Print to Printer' if output_flag == FLAG_PRINT else 'Download to PC'}\n")
                self.log_text.insert(tk.END, "Sending GET_72HRS_DATA command...\n")
                self.log_text.see(tk.END)
                
                # Create the flagged command
                flagged_cmd = f"{output_flag}:B=GET_72HRS_DATA\n".encode('utf-8')
                ser.write(flagged_cmd)
                
                # Variable to store received data
                raw_data = []
                progress = 0
                time_waited = 0
                last_data_time = time.time()
                
                # Read until timeout or specific end marker
                while True:
                    if ser.in_waiting:
                        line = ser.readline().decode('utf-8', errors='ignore').strip()
                        if line:
                            raw_data.append(line)
                            progress += 1
                            last_data_time = time.time()
                            
                            # Update progress (assuming we expect around 1000 lines max)
                            if progress % 10 == 0:
                                self.progress_var.set(min(99, progress / 10))
                                self.log_text.insert(tk.END, f"Received {progress} lines of data...\n")
                                self.log_text.see(tk.END)
                                
                            # Check for end marker if the device sends one
                            if line == "END_OF_DATA" or line == "END" or line == "DOWNLOAD_COMPLETE":
                                break
                    else:
                        # If no data received for a while, consider the download complete
                        time.sleep(0.1)
                        time_waited += 0.1
                        
                        if time.time() - last_data_time > 5.0:  # 5 seconds of no data
                            break
                        
                        if time_waited > 30.0:  # 30 seconds total timeout
                            break
                
                # Process the raw data
                self.process_log_data(raw_data)
                
                # Complete progress bar
                self.progress_var.set(100)
                
                # Log completion
                self.log_text.insert(tk.END, f"\nDownload complete. Received {len(raw_data)} lines of data.\n")
                self.log_text.see(tk.END)
                
        except serial.SerialException as e:
            self.log_text.insert(tk.END, f"Download error: {e}\n")
            self.log_text.see(tk.END)
            messagebox.showerror("Download Error", f"Error during download: {e}")
        except Exception as e:
            self.log_text.insert(tk.END, f"Unexpected error: {e}\n")
            self.log_text.see(tk.END)
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")
        finally:
            self.is_downloading = False
            
    def process_log_data(self, raw_data, is_head_command=False):
        """Process the raw log data into a structured format"""
        try:
            # For all command responses, we don't want to append to existing data
            # The log_data array is now cleared before each command is sent
            
            # Example of parsing log data (adjust based on actual format)
            for line in raw_data:
                # Skip empty lines or headers
                if not line or line.startswith("---") or line.startswith("HEAD"):
                    continue
                        
                # Try to parse line (format may vary depending on device)
                try:
                    # Assuming CSV-like format: DATE,TIME,SPEED,GPS_LAT,GPS_LON,OTHER...
                    parts = line.split(',')
                    
                    if len(parts) >= 2:
                        entry = {
                            'raw_line': line,
                            'parsed': True
                        }
                        
                        # Try to extract date/time (format may vary)
                        if len(parts) >= 2:
                            entry['date'] = parts[0].strip() if parts[0] else "Unknown"
                            entry['time'] = parts[1].strip() if parts[1] else "Unknown"
                            
                        # Try to extract speed if available
                        if len(parts) >= 3:
                            speed_str = parts[2].strip()
                            try:
                                entry['speed'] = float(speed_str) if speed_str else 0.0
                            except ValueError:
                                entry['speed'] = 0.0
                                
                        # Add more fields as needed
                        
                        # Add a flag if this is from a HEAD command
                        entry['is_head_data'] = is_head_command
                        
                        self.log_data.append(entry)
                    else:
                        # Line doesn't match expected format, store as unparsed
                        self.log_data.append({
                            'raw_line': line,
                            'parsed': False,
                            'is_head_data': is_head_command
                        })
                        
                except Exception as e:
                    # If parsing fails, store the raw line
                    self.log_data.append({
                        'raw_line': line,
                        'parsed': False,
                        'error': str(e),
                        'is_head_data': is_head_command
                    })
                    
            # Display summary in the log text if this is a full download (not a command response)
            if self.is_downloading:
                self.log_text.insert(tk.END, f"\nProcessed {len(self.log_data)} log entries.\n")
                
                # Display a few sample entries if available
                if self.log_data:
                    self.log_text.insert(tk.END, "\nSample log entries:\n")
                    for i in range(min(5, len(self.log_data))):
                        self.log_text.insert(tk.END, f"{self.log_data[i]['raw_line']}\n")
                        
        except Exception as e:
            self.log_text.insert(tk.END, f"Error processing log data: {e}\n")
            
    def apply_filter(self):
        """Apply date/time filters to the log data"""
        if not self.log_data:
            messagebox.showinfo("No Data", "No log data available to filter.")
            return
            
        filter_date = self.filter_date_entry.get().strip()
        filter_time = self.filter_time_entry.get().strip()
        
        if not filter_date and not filter_time:
            messagebox.showinfo("No Filter", "Please enter a date and/or time filter.")
            return
            
        filtered_data = []
        
        for entry in self.log_data:
            if not entry.get('parsed', False):
                continue
                
            match = True
            
            # Check date filter if provided
            if filter_date and 'date' in entry:
                if filter_date not in entry['date']:
                    match = False
                    
            # Check time filter if provided
            if filter_time and 'time' in entry:
                if filter_time not in entry['time']:
                    match = False
                    
            if match:
                filtered_data.append(entry)
                
        # Display results
        if filtered_data:
            self.log_text.insert(tk.END, f"\n--- Filter Results ({len(filtered_data)} entries) ---\n")
            
            for entry in filtered_data:
                self.log_text.insert(tk.END, f"{entry['raw_line']}\n")
                
            self.log_text.see(tk.END)
        else:
            self.log_text.insert(tk.END, "\n--- No entries match the filter criteria ---\n")
            self.log_text.see(tk.END)
            
    def plot_speed_vs_time(self):
        """Plot speed vs. time for the log data"""
        if not self.log_data:
            messagebox.showinfo("No Data", "No log data available to plot.")
            return
            
        # Collect the data for plotting
        speeds = []
        times = []
        
        for entry in self.log_data:
            if not entry.get('parsed', False):
                continue
                
            if 'speed' in entry and 'time' in entry:
                speeds.append(entry['speed'])
                times.append(entry['time'])
                
        if not speeds:
            messagebox.showinfo("No Data", "No valid speed/time data found for plotting.")
            return
            
        # Create a new top-level window for the plot
        plot_window = tk.Toplevel(self.root)
        plot_window.title("Speed vs Time Plot")
        plot_window.geometry("800x600")
        
        # Create figure and plot
        fig = Figure(figsize=(8, 6), dpi=100)
        plot = fig.add_subplot(111)
        
        plot.plot(range(len(speeds)), speeds, 'b-')  # Use simple indices for x-axis
        plot.set_title('Speed vs Time')
        plot.set_xlabel('Time Points')
        plot.set_ylabel('Speed (km/h)')
        plot.grid(True)
        
        # Add the time labels at regular intervals
        if len(times) > 10:
            step = len(times) // 10
            x_ticks = range(0, len(times), step)
            x_labels = [times[i] for i in x_ticks]
            plot.set_xticks(x_ticks)
            plot.set_xticklabels(x_labels, rotation=45)
        else:
            plot.set_xticks(range(len(times)))
            plot.set_xticklabels(times, rotation=45)
            
        # Embed the plot in tkinter window
        canvas = FigureCanvasTkAgg(fig, master=plot_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # Add navigation toolbar
        toolbar = NavigationToolbar2Tk(canvas, plot_window)
        toolbar.update()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
    def export_csv(self):
        """Export log data to a CSV file"""
        if not self.log_data:
            messagebox.showinfo("No Data", "No log data available to export.")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
            title="Export Log Data as CSV"
        )
        
        if not file_path:
            return
            
        try:
            # Define field names for HEAD command data
            field_names = [
                "a) VEHICLE OWNERS NAME",
                "b) VEHICLE OWNER ID",
                "c) OWNER PHONE NO",
                "d) VEHICLE REG. NUMBER",
                "e) CHASIS NO",
                "f) VEHICLE MAKE & TYPE",
                "g) CERTIFICATE NUMBER",
                "h) LIMITER TYPE",
                "i) LIMITER SERIAL",
                "j) DATE OF FITTING",
                "k) FITTING AGENTS NAME",
                "l) FITING AGENT ID",
                "m) NAME & LOCATION OF STATION",
                "n) EMAIL ADDRESS",
                "o) AGENT PHONE NUMBER",
                "p) BUSINESS REG. NO"
            ]
            
            # Check if any entry is from HEAD command
            has_head_data = any(entry.get('is_head_data', False) for entry in self.log_data)
            
            with open(file_path, 'w', newline='') as csvfile:
                if has_head_data:
                    # Filter out only HEAD command data
                    head_data = [entry for entry in self.log_data if entry.get('is_head_data', False)]
                    
                    # Set up CSV writer
                    writer = csv.writer(csvfile)
                    
                    # Write field names and values for HEAD data
                    entries_to_process = min(len(field_names), len(head_data))
                    
                    for i in range(entries_to_process):
                        # Write field name and value as a row
                        writer.writerow([field_names[i], head_data[i].get('raw_line', '')])
                else:
                    # For non-HEAD data, use the original export format
                    # Determine fields based on the first parsed entry
                    fields = ['date', 'time', 'speed', 'raw_line']
                    
                    for entry in self.log_data:
                        if entry.get('parsed', False):
                            fields = list(set(fields) | set(entry.keys()))
                            fields = [f for f in fields if f not in ['parsed', 'is_head_data', 'error']]
                            break
                            
                    writer = csv.DictWriter(csvfile, fieldnames=fields, extrasaction='ignore')
                    writer.writeheader()
                    
                    for entry in self.log_data:
                        if entry.get('parsed', False):
                            # Remove unwanted fields
                            data = {k: v for k, v in entry.items() if k in fields}
                            writer.writerow(data)
                        else:
                            # For unparsed entries, write the raw line
                            row = {field: '' for field in fields}
                            row['raw_line'] = entry.get('raw_line', '')
                            writer.writerow(row)
                            
            messagebox.showinfo("Export Successful", f"Log data exported to:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export CSV file: {e}")
            
    def export_excel(self):
        """Export log data to an Excel file"""
        if not self.log_data:
            messagebox.showinfo("No Data", "No log data available to export.")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            title="Export Log Data as Excel"
        )
        
        if not file_path:
            return
            
        try:
            # Create a new workbook and select active worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Log Data"
            
            # Define field names for HEAD command data
            field_names = [
                "a) VEHICLE OWNERS NAME",
                "b) VEHICLE OWNER ID",
                "c) OWNER PHONE NO",
                "d) VEHICLE REG. NUMBER",
                "e) CHASIS NO",
                "f) VEHICLE MAKE & TYPE",
                "g) CERTIFICATE NUMBER",
                "h) LIMITER TYPE",
                "i) LIMITER SERIAL",
                "j) DATE OF FITTING",
                "k) FITTING AGENTS NAME",
                "l) FITING AGENT ID",
                "m) NAME & LOCATION OF STATION",
                "n) EMAIL ADDRESS",
                "o) AGENT PHONE NUMBER",
                "p) BUSINESS REG. NO"
            ]
            
            # Check if any entry is from HEAD command
            has_head_data = any(entry.get('is_head_data', False) for entry in self.log_data)
            
            if has_head_data:
                # Filter out only HEAD command data
                head_data = [entry for entry in self.log_data if entry.get('is_head_data', False)]
                
                # Write field names and values for HEAD data
                entries_to_process = min(len(field_names), len(head_data))
                
                for row_num, i in enumerate(range(entries_to_process), 1):
                    # Field name in column A
                    ws.cell(row=row_num, column=1).value = field_names[i]
                    
                    # Value in column B
                    ws.cell(row=row_num, column=2).value = head_data[i].get('raw_line', '')
            else:
                # For non-HEAD data, use the original export format
                # Determine fields based on the first parsed entry
                fields = ['date', 'time', 'speed', 'raw_line']
                
                for entry in self.log_data:
                    if entry.get('parsed', False):
                        fields = list(set(fields) | set(entry.keys()))
                        fields = [f for f in fields if f not in ['parsed', 'is_head_data', 'error']]
                        break
                        
                # Write header row
                for col_num, field in enumerate(fields, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = field
                    
                # Write data rows
                for row_num, entry in enumerate(self.log_data, 2):
                    if entry.get('parsed', False):
                        for col_num, field in enumerate(fields, 1):
                            cell = ws.cell(row=row_num, column=col_num)
                            cell.value = entry.get(field, '')
                    else:
                        # For unparsed entries, write the raw line
                        raw_line_col = fields.index('raw_line') + 1 if 'raw_line' in fields else 1
                        cell = ws.cell(row=row_num, column=raw_line_col)
                        cell.value = entry.get('raw_line', '')
                        
            # Save the workbook
            wb.save(file_path)
            messagebox.showinfo("Export Successful", f"Log data exported to:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export Excel file: {e}")
            
    def export_pdf(self):
        """Export log data to a PDF file"""
        if not self.log_data:
            messagebox.showinfo("No Data", "No log data available to export.")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")],
            title="Export Log Data as PDF"
        )
        
        if not file_path:
            return
            
        try:
            # Create a new PDF document
            c = canvas.Canvas(file_path, pagesize=letter)
            width, height = letter
            
            # Set title
            c.setFont("Helvetica-Bold", 16)
            c.drawString(inch, height - inch, "Vehicle Limiter Log Data")
            
            # Add timestamp
            c.setFont("Helvetica", 10)
            c.drawString(inch, height - 1.25*inch, 
                           f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            # Add header for data
            c.setFont("Helvetica-Bold", 12)
            y_position = height - 2*inch
            c.drawString(inch, y_position, "Log Data Entries:")
            y_position -= 0.25*inch
            
            # Draw line under header
            c.line(inch, y_position, width - inch, y_position)
            y_position -= 0.5*inch
            
            # Define field names for HEAD command data
            field_names = [
                "a) VEHICLE OWNERS NAME",
                "b) VEHICLE OWNER ID",
                "c) OWNER PHONE NO",
                "d) VEHICLE REG. NUMBER",
                "e) CHASIS NO",
                "f) VEHICLE MAKE & TYPE",
                "g) CERTIFICATE NUMBER",
                "h) LIMITER TYPE",
                "i) LIMITER SERIAL",
                "j) DATE OF FITTING",
                "k) FITTING AGENTS NAME",
                "l) FITING AGENT ID",
                "m) NAME & LOCATION OF STATION",
                "n) EMAIL ADDRESS",
                "o) AGENT PHONE NUMBER",
                "p) BUSINESS REG. NO"
            ]
            
            # First, check if any entry is from HEAD command
            has_head_data = any(entry.get('is_head_data', False) for entry in self.log_data)
            
            # Set font for data
            c.setFont("Courier", 10)
            
            if has_head_data:
                # Filter out only HEAD command data
                head_data = [entry for entry in self.log_data if entry.get('is_head_data', False)]
                
                # Process HEAD data with field names
                entries_to_process = min(len(field_names), len(head_data))
                
                for i in range(entries_to_process):
                    # Skip if not enough space on current page
                    if y_position < 2*inch:
                        c.showPage()
                        y_position = height - inch
                        c.setFont("Courier", 10)
                    
                    # Get field name and value
                    field_name = field_names[i]
                    line = head_data[i].get('raw_line', '')
                    
                    # Format with dots for better readability
                    dots = "." * (30 - len(field_name))
                    formatted_line = f"{field_name}{dots} {line}"
                    
                    # Write formatted entry
                    c.drawString(inch, y_position, formatted_line[:100])
                    
                    # If line is longer than 100 chars, add continuation on next line
                    if len(formatted_line) > 100:
                        y_position -= 0.2*inch
                        c.drawString(1.2*inch, y_position, formatted_line[100:200])
                    
                    y_position -= 0.3*inch
            else:
                # For non-HEAD data, just print raw lines
                for entry in self.log_data:
                    # Skip if not enough space on current page
                    if y_position < 2*inch:
                        c.showPage()
                        y_position = height - inch
                        c.setFont("Courier", 10)
                    
                    # Write log entry
                    line = entry.get('raw_line', '')
                    c.drawString(inch, y_position, line[:100])  # Limit line length
                    
                    # If line is longer than 100 chars, add continuation on next line
                    if len(line) > 100:
                        y_position -= 0.2*inch
                        c.drawString(1.2*inch, y_position, line[100:200])
                    
                    y_position -= 0.3*inch
                
            # Save the PDF
            c.save()
            messagebox.showinfo("Export Successful", f"Log data exported to:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export PDF file: {e}")
            
    def clear_log(self):
        """Clear the log text widget"""
        self.log_text.delete(1.0, tk.END)
        
    def save_log(self):
        """Save the log contents to a text file"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")],
            title="Save Log Output"
        )
        
        if file_path:
            try:
                with open(file_path, 'w') as f:
                    log_content = self.log_text.get(1.0, tk.END)
                    f.write(log_content)
                messagebox.showinfo("Save Successful", f"Log saved to:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Save Error", f"Could not save log: {e}")

# ----- Main Application -----
if __name__ == "__main__":
    root = tk.Tk()
    root.title("Vehicle Limiter Control & Log Downloader")
    root.geometry("900x800")  # More space for the expanded UI
    
    app = LogDownloaderApp(root)
    
    root.mainloop()


